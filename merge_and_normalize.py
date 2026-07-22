"""
Welder Qualification Merge & Normalize Tool
=============================================

Purpose
-------
Takes two files:
  1. CLIENT LIST  - a fresh export with 'SCG' and 'Contractor' tabs, NO IDs.
  2. PROD FILE    - the current PROD reference export, which has an 'Append'
                     tab containing an ID column (col A) and Employee Number
                     (col B). This is used ONLY to look up IDs by Employee
                     Number - Employee Number is the bridge between the two
                     files.

What it does
------------
  1. Combines the client list's SCG + Contractor tabs into one unified table
     (mirrors the manual "Append" logic: Location Name is populated for SCG
     rows, left blank for Contractor rows since Contractor tracks
     "Contracting Company Name" instead, which is dropped; Job Name (SCG)
     and Contracting Company Name (Contractor) are both dropped).
  2. Builds an ID lookup {Employee Number -> ID} from the PROD file's
     Append tab.
  3. Groups the client list by Employee Number and collapses duplicate
     qualification rows per employee:
       - Qualifications are concatenated as "QualName (Exp: M/D/YYYY)",
         joined with " & ", ordered oldest-expiration-first.
       - Current Qualification Type Name, Next Qualification Type Name,
         and the two date columns come from the row with the LATEST End Date.
       - Status comes from the row with the EARLIEST End Date.
  4. Attaches the ID from the PROD lookup. Employees not found in PROD are
     marked "Not Found" (to be assigned manually).
  5. Writes a normalized output workbook, ready to use for a PROD DB update.

Usage
-----
    python3 merge_and_normalize.py <client_list.xlsx> <prod_file.xlsx> <output.xlsx>

Example
-------
    python3 merge_and_normalize.py \\
        /mnt/user-data/uploads/Client_List.xlsx \\
        /mnt/user-data/uploads/Client_List_with_FM_IDs.xlsx \\
        /mnt/user-data/outputs/Client_List_Normalized.xlsx
"""

import sys
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def fmt_date(d):
    if d is None:
        return ''
    return f"{d.month}/{d.day}/{d.year}"


def combine_client_list(client_path):
    """Combine the SCG + Contractor tabs into one unified row list.

    Unified columns: Employee Number, Employee Name, Qualification Name,
    Begin Date, End Date, Current Qualification Type Name,
    Next Qualification Type Name, Next Qualification Date, Location Name,
    Status
    """
    wb = openpyxl.load_workbook(client_path, data_only=True)

    if 'SCG' not in wb.sheetnames or 'Contractor' not in wb.sheetnames:
        raise ValueError(
            f"Expected 'SCG' and 'Contractor' tabs in {client_path}, "
            f"found: {wb.sheetnames}"
        )

    combined = []

    scg_rows = list(wb['SCG'].iter_rows(min_row=2, values_only=True))
    for r in scg_rows:
        # SCG: EmpNum, Name, Qual, Begin, End, CurType, NextType, NextDate, Location, Status, JobName(dropped)
        combined.append((
            str(r[0]), r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9]
        ))

    con_rows = list(wb['Contractor'].iter_rows(min_row=2, values_only=True))
    for r in con_rows:
        # Contractor: EmpNum, Name, Qual, Begin, End, CurType, NextType, NextDate, ContractingCo(dropped), Status
        combined.append((
            str(r[0]), r[1], r[2], r[3], r[4], r[5], r[6], r[7], '', r[9]
        ))

    return combined


def build_id_lookup(prod_path):
    """Build {Employee Number -> ID} from the PROD file's Append tab."""
    wb = openpyxl.load_workbook(prod_path, data_only=True)

    if 'Append' not in wb.sheetnames:
        raise ValueError(
            f"Expected an 'Append' tab with ID + Employee Number in "
            f"{prod_path}, found: {wb.sheetnames}"
        )

    ws = wb['Append']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # Append tab columns: ID, Employee Number, Employee Name, Qualification...

    lookup = {}
    for r in rows:
        emp_id, emp_num = r[0], str(r[1])
        if emp_num not in lookup or lookup[emp_num] == 'Not Found':
            lookup[emp_num] = emp_id

    return lookup


def normalize(combined_rows, id_lookup):
    groups = defaultdict(list)
    for r in combined_rows:
        groups[r[0]].append(r)

    output_rows = []
    for emp_num, recs in groups.items():
        name = recs[0][1]
        emp_id = id_lookup.get(emp_num, 'Not Found')

        recs_sorted = sorted(recs, key=lambda r: r[4])  # End Date ascending
        concat_parts = [f"{r[2]} (Exp: {fmt_date(r[4])})" for r in recs_sorted]
        concat_qual = " & ".join(concat_parts)

        latest_rec = max(recs, key=lambda r: r[4])
        earliest_rec = min(recs, key=lambda r: r[4])

        latest_end = latest_rec[4]
        latest_begin = latest_rec[3]
        cur_qual_type = latest_rec[5]
        next_qual_type = latest_rec[6]
        status = earliest_rec[9]

        output_rows.append([
            emp_id,
            emp_num,
            name,
            fmt_date(latest_end),
            fmt_date(latest_begin),
            concat_qual,
            cur_qual_type,
            next_qual_type,
            status,
        ])

    output_rows.sort(key=lambda r: r[1])
    return output_rows


def write_output(output_rows, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Normalized"

    headers = [
        "UUID",
        "Employee Number",
        "Employee Name",
        "Latest Expiration Date",
        "Begin Date (Latest Qual)",
        "Qualifications",
        "Current Qualification Type Name",
        "Next Qualification Type Name",
        "Status",
    ]
    ws.append(headers)

    font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for row in output_rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = font

    widths = [38, 16, 26, 20, 20, 70, 26, 26, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)


def main():
    if len(sys.argv) != 4:
        print("Usage: python3 merge_and_normalize.py <client_list.xlsx> <prod_file.xlsx> <output.xlsx>")
        sys.exit(1)

    client_path, prod_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3]

    combined_rows = combine_client_list(client_path)
    id_lookup = build_id_lookup(prod_path)
    output_rows = normalize(combined_rows, id_lookup)
    write_output(output_rows, out_path)

    not_found = [r[1] for r in output_rows if r[0] == 'Not Found']
    print(f"Unique employees: {len(output_rows)}")
    print(f"Total source rows combined: {len(combined_rows)}")
    print(f"Employees with 'Not Found' ID: {len(not_found)}")
    if not_found:
        print("  ", not_found)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
